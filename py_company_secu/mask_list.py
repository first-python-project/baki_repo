from flask import Flask, send_file
import openpyxl
import re

app = Flask(__name__)

@app.route('/')
def list():
    file_path = "custom_File\customer_list.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    
    #파일 내 중요 정보(전화번호 가운데 3, 4자리, 주소에서 도,시,구로 끝나는 것 제외, 이메일) 정규표현식
    phone_number_pattern = r"\d{2,3}-\d{3,4}-\d{4}"      #전화번호 가운데 4자리 찾기
    email_pattern = "[\w\.-]+@[\w\.-]+"                 #이메일 찾기

    # 모든 시트 순회
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 각 셀 순회하며 중요정보 찾기
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    # 셀의 값에서 중요정보 각 패턴 찾기
                    phone_number_in_cell = re.findall(phone_number_pattern, str(cell.value))
                    email_in_cell = re.findall(email_pattern, str(cell.value))

                    if cell.column == 1:  # 이름이 있는 열
                        name = cell.value
                        if len(name) > 1:  # 이름의 길이가 2자 이상인 경우 (성이 포함된 경우)
                            masked_name = name[0] + '*' * (len(name) - 1)  # 성을 제외한 나머지 부분을 마스킹 처리
                            cell.value = masked_name

                    if phone_number_in_cell:
                        # 전화번호 가운데 자릿수가 3자리, 4자리인 경우도 마스킹 처리
                        for phone_number in phone_number_in_cell:
                            split_phone_number = phone_number.split('-')
                            if len(split_phone_number[1]) == 3:
                                #전화번호 가운데 3자리 마스킹 처리
                                masked_phone_number = phone_number[:3] + "-***-" + phone_number[-4:]
                            else:
                                #전화번호 가운데 4자리 마스킹 처리
                                masked_phone_number = phone_number[:4] + "-****-" + phone_number[-4:]
                        cell.value = cell.value.replace(phone_number, masked_phone_number)
                    if email_in_cell:
                        for email in email_in_cell:
                            # 이메일을 마스킹 처리
                            masked_email = email[:4] + "*****" + email[email.index("@"):]
                            cell.value = cell.value.replace(email, masked_email)

    # 수정된 내용을 새로운 파일로 저장
    masked_file_path = "custom_File\masked_your_excel_file.xlsx"
    workbook.save(masked_file_path)
    
    # 클라이언트에게 다운로드할 파일을 응답으로 전달
    return send_file(masked_file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)