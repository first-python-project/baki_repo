import openpyxl
from faker import Faker

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = "이름"
worksheet['B1'] = "전화번호"
worksheet['E1'] = "이메일"
worksheet['F1'] = "IP주소"

fake = Faker('ko_KR')


for row in range(2, 50): #2행부터 시작해 50줄까지 생성
    worksheet.cell(row=row, column=1, value=fake.name())
    worksheet.cell(row=row, column=2, value=fake.phone_number())
    worksheet.cell(row=row, column=3, value=fake.postcode())
    worksheet.cell(row=row, column=4, value=fake.address())
    worksheet.cell(row=row, column=5, value=fake.email())
    worksheet.cell(row=row, column=6, value=fake.ipv4_private())

workbook.save("customer_list.xlsx")


